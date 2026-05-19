from flask import Flask, request
from openpyxl import load_workbook
from datetime import datetime
import requests

app = Flask(__name__)

EXCEL_FILE = "eligibility.xlsx"

VERIFY_TOKEN = "rfd123"
ACCESS_TOKEN = "EAASb7CLVZBnEBRQLXesYI9s91n8ZAI3HFUg2twVUHZAV7qzRQUpmHfB02J6z3xHg7rykox3aifcj9XJbEUKdhC6fjNoJMfoEpaJ2boUiXjU53nZBfIx6WD3PqNEovZCccU0o82GcY2N8sZCWswHsuQoDLBEfjyi5ZBW7e0XmWVVqFGOlZAUWyK8HJrr6Q20GMoOEZC7hMFkn8m55VcQAtRQKDNXXwy0Rb1lygqOkCiLYKxgZCRw00F5ZCe7HFAebPWmfryirrjMSOAZBtq4czBcZBZAzQI3E7I"
PHONE_NUMBER_ID = "1190519857470878"

SHEET_RESPONSE = "Response Master"
SHEET_JANMAR = "Reg. CASAvaganza Jan-Mar"
SHEET_APRMEI = "Reg. CASAvaganza 2.0 (Apr-Mei)"


def normalisasi(value):
    return str(value or "").strip().upper().replace(" ", "").replace("\n", "").replace("\t", "")


def send_wa(to, message):
    url = f"https://graph.facebook.com/v25.0/{PHONE_NUMBER_ID}/messages"

    headers = {
        "Authorization": f"Bearer {ACCESS_TOKEN}",
        "Content-Type": "application/json"
    }

    payload = {
        "messaging_product": "whatsapp",
        "to": to,
        "type": "text",
        "text": {"body": message}
    }

    response = requests.post(url, headers=headers, json=payload)
    print("SEND WA:", response.status_code, response.text)


def cari_cif_di_response_master(cif_input):
    wb = load_workbook(EXCEL_FILE, data_only=True)
    ws = wb[SHEET_RESPONSE]
    cif_input = normalisasi(cif_input)

    for row in range(2, ws.max_row + 1):
        cif = normalisasi(ws[f"J{row}"].value)
        if cif == cif_input:
            return True

    return False


def cek_dan_kirim_voucher(cif_input, sheet_name, nomor_wa):
    wb = load_workbook(EXCEL_FILE)
    ws = wb[sheet_name]
    cif_input = normalisasi(cif_input)

    for row in range(2, ws.max_row + 1):
        cif = normalisasi(ws[f"G{row}"].value)

        if cif == cif_input:
            nama = str(ws[f"H{row}"].value or "").strip()
            keterangan = str(ws[f"W{row}"].value or "").strip()
            status_bot = str(ws[f"AF{row}"].value or "").strip()

            if keterangan:
                return f"Belum eligible.\n\nKeterangan:\n{keterangan}"

            if status_bot:
                return (
                    "Voucher untuk CIF tersebut sudah pernah dikirim melalui Hotline RFD.\n\n"
                    f"Status:\n{status_bot}"
                )

            vouchers = []
            for col in ["AC", "AD", "AE"]:
                v = str(ws[f"{col}{row}"].value or "").strip()
                if v:
                    vouchers.append(v)

            if not vouchers:
                return "Data eligible, namun kode voucher belum tersedia. Mohon hubungi admin Hotline RFD."

            waktu_kirim = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws[f"AF{row}"] = f"TERKIRIM VIA BOT | {waktu_kirim} | {nomor_wa}"
            wb.save(EXCEL_FILE)

            voucher_text = "\n".join([f"- {v}" for v in vouchers])

            return (
                f"Selamat {nama} 🎉\n\n"
                "Berdasarkan hasil pengecekan, Anda memenuhi ketentuan Program CASAvaganza.\n\n"
                f"Kode voucher Anda:\n{voucher_text}\n\n"
                "Mohon simpan kode ini. Voucher hanya dikirimkan 1 kali untuk setiap CIF sesuai ketentuan program."
            )

    return f"CIF tidak ditemukan di sheet {sheet_name}."


def proses_cif(cif_input, nomor_wa):
    if not cari_cif_di_response_master(cif_input):
        return "CIF tidak ditemukan pada Response Master."

    hasil_janmar = cek_dan_kirim_voucher(cif_input, SHEET_JANMAR, nomor_wa)

    if "tidak ditemukan di sheet" not in hasil_janmar.lower():
        return hasil_janmar

    hasil_aprmei = cek_dan_kirim_voucher(cif_input, SHEET_APRMEI, nomor_wa)

    if "tidak ditemukan di sheet" not in hasil_aprmei.lower():
        return hasil_aprmei

    return "CIF ditemukan di Response Master, namun tidak ditemukan di data registrasi CASAvaganza."


@app.route("/webhook", methods=["GET"])
def verify_webhook():
    mode = request.args.get("hub.mode")
    token = request.args.get("hub.verify_token")
    challenge = request.args.get("hub.challenge")

    if mode == "subscribe" and token == VERIFY_TOKEN:
        return challenge, 200

    return "Forbidden", 403


@app.route("/webhook", methods=["POST"])
def receive_message():
    data = request.get_json()
    print("INCOMING:", data)

    try:
        message = data["entry"][0]["changes"][0]["value"]["messages"][0]
        sender = message["from"]
        text = message["text"]["body"].strip()

        if not any(char.isdigit() for char in text):
            reply = (
                "Halo, terima kasih telah menghubungi Hotline RFD.\n\n"
                "Silakan masukkan CIF untuk pengecekan Program CASAvaganza."
            )
            send_wa(sender, reply)
            return "OK", 200
        
        cif_input = normalisasi(text)

        reply = proses_cif(cif_input, sender)
        send_wa(sender, reply)

    except Exception as e:
        print("ERROR:", e)

    return "OK", 200


if __name__ == "__main__":
    app.run(port=5000, debug=True)
